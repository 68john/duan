# QL_Diem_HS (Upgraded UI Edition, AI)
# -*- coding: utf-8 -*-
from __future__ import annotations
from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    import pandas as _pd_check  # type: ignore

import os, sys, csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Biểu đồ
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import numpy as np

# Load .env if present
try:
    from dotenv import load_dotenv, find_dotenv
    def _load_dotenv_robust():
        base_dir = os.path.dirname(sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__))
        candidates = [
            os.path.join(base_dir, ".env"),
            find_dotenv(usecwd=True),
        ]
        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            candidates.insert(1, os.path.join(sys._MEIPASS, ".env"))
        for path in candidates:
            if path and os.path.isfile(path):
                load_dotenv(path, override=False)
                break
    _load_dotenv_robust()
except Exception:
    pass

# (tuỳ chọn) Logo
try:
    from PIL import Image, ImageTk
    _HAS_PIL = True
except Exception:
    _HAS_PIL = False

# ---- Cấu hình font Unicode cho Matplotlib ----
matplotlib.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Segoe UI']
matplotlib.rcParams['axes.unicode_minus'] = False

def resource_path(rel_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        base = sys._MEIPASS
    else:
        base = os.path.abspath(".")
    return os.path.join(base, rel_path)

SUBJECTS = ["toan", "ly", "hoa", "van", "anh", "tin"]

# ================== MẢNG THỐNG KÊ (10 BIN) & MẢNG RAW ==================
# 10 bin: [0,1), [1,2), ..., [8,9), [9,10]
TOAN = [0]*10; LI = [0]*10; HOA = [0]*10
VAN  = [0]*10; ANH = [0]*10; TIN = [0]*10; TB = [0]*10

TOAN_RAW, LI_RAW, HOA_RAW = [], [], []
VAN_RAW, ANH_RAW, TIN_RAW = [], [], []
TB_RAW = []

# ================== HÀM TÍNH TOÁN ==================
def classify(avg: float) -> str:
    if avg >= 8.5: return "Giỏi"
    if avg >= 7.0: return "Khá"
    if avg >= 5.0: return "Trung bình"
    return "Yếu"

def wavg(scores: dict) -> float:
    return sum(scores.values())/len(scores) if scores else 0.0

def parse_score_any(txt: str) -> float:
    if txt is None: return 0.0
    s = str(txt).strip()
    if not s: return 0.0
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        v = float(s)
        if 0 <= v <= 10: return v
    except:
        pass
    return 0.0

def _bucket_index(v: float) -> int:
    try:
        x = float(v)
    except:
        x = 0.0
    if x < 0: x = 0.0
    if x > 10: x = 10.0
    return 9 if x >= 9.0 else int(x)

def update_histograms_from_raw():
    global TOAN, LI, HOA, VAN, ANH, TIN, TB
    TOAN = [0]*10; LI = [0]*10; HOA = [0]*10
    VAN  = [0]*10; ANH = [0]*10; TIN = [0]*10; TB = [0]*10
    for s in TOAN_RAW: TOAN[_bucket_index(s)] += 1
    for s in LI_RAW:   LI[_bucket_index(s)]   += 1
    for s in HOA_RAW:  HOA[_bucket_index(s)]  += 1
    for s in VAN_RAW:  VAN[_bucket_index(s)]  += 1
    for s in ANH_RAW:  ANH[_bucket_index(s)]  += 1
    for s in TIN_RAW:  TIN[_bucket_index(s)]  += 1
    for s in TB_RAW:   TB[_bucket_index(s)]   += 1

# ---------- AI helpers (để giữ tương thích với nền cũ) ----------
def _ai__summarize_records(records, limit_rows=15):
    if not records:
        return "Bảng hiện đang trống."
    ranks = {}
    for s in records:
        xl = (s.get("xep_loai") if isinstance(s, dict) else getattr(s, "xep_loai", "")) or \
             (s.get("xếp loại") if isinstance(s, dict) else "")
        ranks[xl] = ranks.get(xl, 0) + 1
    head = []
    for i, s in enumerate(records[:limit_rows], start=1):
        if isinstance(s, dict):
            name = s.get("ho_ten") or s.get("Họ Tên") or "?"
            clss = s.get("lop") or s.get("Lớp") or "?"
            tb = s.get("diem_tb") or s.get("Điểm TB") or 0.0
        else:
            name = getattr(s, "ho_ten", "?"); clss = getattr(s, "lop", "?"); tb = getattr(s, "diem_tb", 0.0)
        try: tb = float(tb)
        except Exception: tb = 0.0
        xl = (s.get("xep_loai") if isinstance(s, dict) else getattr(s, "xep_loai","?")) or s.get("xếp loại") if isinstance(s, dict) else "?"
        head.append(f"{i}. {name} ({clss}) - TB={tb:.2f} - XL={xl}")
    return "\n".join([
        f"Số HS hiển thị: {len(records)}",
        "Phân bố xếp loại: " + ", ".join([f"{k}={v}" for k,v in ranks.items()]) if ranks else "(không có)",
        "Một số dòng đầu:",
        *head
    ])

def _ai__ask_chatgpt(question: str, context: str = "", temperature: float = 0.2, max_output_tokens: int = 400) -> str:
    try:
        from openai import OpenAI
    except Exception:
        return "Chưa cài 'openai'. Hãy chạy: pip install --upgrade openai python-dotenv"
    import os

    api_key = os.getenv("OPENAI_API_KEY")
    model   = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    org     = os.getenv("OPENAI_ORG", None)  # optional

    if not api_key:
        return "Chưa thiết lập OPENAI_API_KEY. Đặt biến môi trường hoặc file .env."

    candidates = [model, "gpt-4o-mini", "gpt-4o"]
    client = OpenAI(api_key=api_key, organization=org) if org else OpenAI(api_key=api_key)

    system_prompt = (
        "Bạn là trợ lý AI cho phần mềm Quản lý điểm học sinh. "
        "Trả lời ngắn gọn, chính xác, bằng tiếng Việt. "
        "Nếu câu hỏi liên quan đến dữ liệu bảng đang hiển thị, hãy giải thích cách lọc/tìm/xếp hạng trong app."
    )

    last_err = None
    for md in candidates:
        try:
            rsp = client.responses.create(
                model=md,
                input=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user",
                     "content": (question if not context else f"Ngữ cảnh:\n{context}\n\nCâu hỏi: {question}")},
                ],
                max_output_tokens=max_output_tokens,
                temperature=temperature,
            )
            text = getattr(rsp, "output_text", "").strip()
            if text:
                return text
            last_err = "Không nhận được nội dung trả lời."
        except Exception as e:
            low = str(e).lower()
            if ("not found" in low or "does not exist" in low or "permission" in low or "unsupported" in low):
                last_err = f"Model `{md}` không khả dụng, thử model khác..."
                continue
            if "insufficient_quota" in low or "exceeded your current quota" in low:
                return (
                    "API key hiện không còn quota/credit để gọi.\n"
                    "- Billing: https://platform.openai.com/account/billing\n"
                    "- Usage: https://platform.openai.com/usage\n"
                    "- Nếu bạn có nhiều tổ chức (org), hãy đặt OPENAI_ORG cho đúng."
                )
            if "rate" in low and "limit" in low:
                return "Đang chạm rate limit. Hãy thử lại sau vài giây."
            if "invalid_api_key" in low or "401" in low:
                return "API key không hợp lệ hoặc đã bị thu hồi. Tạo key mới và cập nhật OPENAI_API_KEY."
            last_err = f"Lỗi gọi ChatGPT ({md}): {e}"
            break

    return last_err or "Không nhận được nội dung trả lời."

# ================== ỨNG DỤNG GUI ==================
class StudentManagerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Quản lý điểm học sinh")
        self.root.geometry("1280x800")

        # Trạng thái giao diện
        self.dark_mode = False
        self.buttons: list[tuple[tk.Button, str, str, str]] = []  # (btn, bg, fg, active_bg)
        self.sort_state = {}  # cột -> asc/desc
        self.display_columns = ["stt","id","ho_ten","lop"] + SUBJECTS + ["diem_tb","xep_loai"]

        # Dữ liệu
        self.students = []
        self.next_id = 1

        # KPI variables
        self.kpi_total = tk.StringVar(value="0")
        self.kpi_visible = tk.StringVar(value="0")
        self.kpi_gioi = tk.StringVar(value="0")
        self.kpi_kha = tk.StringVar(value="0")
        self.kpi_tb = tk.StringVar(value="0")
        self.kpi_yeu = tk.StringVar(value="0")

        # Theme gốc
        self._setup_theme()

        # UI
        self._build_header()
        self._build_kpis()           # NEW: dải KPI
        self._build_form()
        self._build_buttons()
        self._build_search()
        self._build_table()
        self._build_statusbar()
        self._set_status("Sẵn sàng.")

        # Shortcuts
        self._bind_shortcuts()

    # ---------- THEME ----------
    def _setup_theme(self):
        self.primary = "#1e88e5"
        self.primary_dark = "#1565c0"
        self.bg_light = "#eaf2fb"
        self.text_dark = "#1f2937"

        style = ttk.Style()
        try: style.theme_use("clam")
        except: pass

        self.root.configure(bg=self.bg_light)
        style.configure("TFrame", background=self.bg_light)
        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"),
                        foreground=self.primary, background=self.bg_light)
        style.configure("KPI.TLabel", font=("Segoe UI", 11, "bold"),
                        foreground=self.text_dark, background=self.bg_light)
        style.configure("KPINum.TLabel", font=("Segoe UI", 16, "bold"),
                        foreground=self.primary, background=self.bg_light)
        style.configure("TLabel", background=self.bg_light, foreground=self.text_dark, font=("Segoe UI", 10))
        style.configure("TLabelframe", background=self.bg_light)
        style.configure("TLabelframe.Label", background=self.bg_light, foreground=self.text_dark,
                        font=("Segoe UI", 11, "bold"))
        style.configure("TEntry", fieldbackground="white", foreground=self.text_dark, background="white")
        style.configure("TCombobox", fieldbackground="white", foreground=self.text_dark, background="white")
        style.configure("Treeview", font=("Segoe UI", 10),
                        background="white", fieldbackground="white", foreground="#000000")
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background=self.primary, foreground="white")
        style.map("Treeview.Heading", background=[("active", self.primary_dark)])

    def _apply_dark_palette(self):
        style = ttk.Style()
        bg = "#1e1e1e"; fg = "#f5f5f5"
        
        self.root.configure(bg=bg)
        style.configure("TFrame", background=bg)
        style.configure("TLabelframe", background=bg)
        style.configure("TLabelframe.Label", background=bg, foreground=fg)
        style.configure("TLabel", background=bg, foreground=fg)
        style.configure("Title.TLabel", background=bg, foreground="#90caf9")
        style.configure("KPI.TLabel", background=bg, foreground=fg)
        style.configure("KPINum.TLabel", background=bg, foreground="#90caf9")
        style.configure("TEntry", fieldbackground="#333333", foreground=fg, background="#333333")
        style.configure("TCombobox", fieldbackground="#333333", foreground=fg, background="#333333")
        style.configure("Treeview", background="#2d2d2d", fieldbackground="#2d2d2d", foreground=fg)
        style.configure("Treeview.Heading", background="#3a3a3a", foreground=fg)


        # đồng bộ màu nền cho vùng nút có thanh cuộn
        if hasattr(self, "btn_canvas"):
            try:
                self.btn_canvas.configure(bg=bg)
            except: pass
        if hasattr(self, "btn_frame"):
            try:
                self.btn_frame.configure(bg=bg)
            except: pass
        if hasattr(self, "logo_lbl"):
            try: self.logo_lbl.configure(bg=bg)
            except: pass

        if hasattr(self, "tree"):
            self._configure_row_tags(dark=True)

        for btn, _bg, _fg, _abg in self.buttons:
            try: btn.config(bg="#333333", fg="#f5f5f5", activebackground="#555555")
            except: pass

    def _apply_light_palette(self):
        bg = "#1e1e1e"

        self._setup_theme()
        self.root.configure(bg=self.bg_light)
        for w in self.root.winfo_children():
            try:
                if isinstance(w, (tk.Frame, tk.Label, tk.Button)):
                    w.configure(bg=self.bg_light)
            except:
                pass
            try:
                for c in w.winfo_children():
                    if isinstance(c, (tk.Frame, tk.Label)):
                        c.configure(bg=self.bg_light)
            except:
                pass


        # đồng bộ màu nền cho vùng nút có thanh cuộn
        if hasattr(self, "btn_canvas"):
            try:
                self.btn_canvas.configure(bg=bg)
            except: pass
        if hasattr(self, "btn_frame"):
            try:
                self.btn_frame.configure(bg=bg)
            except: pass
        if hasattr(self, "logo_lbl"):
            try: self.logo_lbl.configure(bg=self.bg_light)
            except: pass

        # đồng bộ màu nền cho vùng nút có thanh cuộn
        if hasattr(self, "btn_canvas"):
            try:
                self.btn_canvas.configure(bg=self.bg_light)
            except: pass
        if hasattr(self, "btn_frame"):
            try:
                self.btn_frame.configure(bg=self.bg_light)
            except: pass


        if hasattr(self, "tree"):
            self._configure_row_tags(dark=False)

        for btn, bgc, fgc, abg in self.buttons:
            try:
                btn.config(bg=bgc, fg=fgc, activebackground=abg)
            except Exception:
                pass

    def _configure_row_tags(self, dark: bool):
        # màu theo xếp loại
        if dark:
            self.tree.tag_configure("rank_gioi",  background="#113d1d", foreground="#e8ffe8")
            self.tree.tag_configure("rank_kha",   background="#0d2d4a", foreground="#e8f4ff")
            self.tree.tag_configure("rank_tb",    background="#453e13", foreground="#fffce8")
            self.tree.tag_configure("rank_yeu",   background="#4a1010", foreground="#ffe8e8")
            self.tree.tag_configure("oddrow",  background="#1f1f1f", foreground="#f5f5f5")
            self.tree.tag_configure("evenrow", background="#262626", foreground="#f5f5f5")
        else:
            self.tree.tag_configure("rank_gioi",  background="#E8F5E9", foreground="#1b5e20")
            self.tree.tag_configure("rank_kha",   background="#E3F2FD", foreground="#0b3060")
            self.tree.tag_configure("rank_tb",    background="#FFF8E1", foreground="#6d4c00")
            self.tree.tag_configure("rank_yeu",   background="#FFEBEE", foreground="#b71c1c")
            self.tree.tag_configure("oddrow",  background="#ffffff", foreground="#000000")
            self.tree.tag_configure("evenrow", background="#f6fbff", foreground="#000000")

    def toggle_dark_mode(self, *_):
        self.dark_mode = not self.dark_mode
        if self.dark_mode:
            self._apply_dark_palette()
        else:
            self._apply_light_palette()
        self._set_status("Đã đổi chế độ giao diện.")

    def _color_picker(self):
        color = colorchooser.askcolor(title="Chọn màu chủ đạo (Primary)")
        if not color or not color[1]:
            return
        self.primary = color[1]
        # làm tối một chút cho active
        self.primary_dark = self.primary
        try:
            import colorsys
            r = int(self.primary[1:3],16)/255
            g = int(self.primary[3:5],16)/255
            b = int(self.primary[5:7],16)/255
            h,l,s = colorsys.rgb_to_hls(r,g,b)
            l = max(0, l-0.15)
            r2,g2,b2 = colorsys.hls_to_rgb(h,l,s)
            self.primary_dark = f"#{int(r2*255):02x}{int(g2*255):02x}{int(b2*255):02x}"
        except Exception:
            pass
        self._setup_theme()
        # cập nhật màu cho tree heading
        style = ttk.Style()
        style.configure("Treeview.Heading", background=self.primary, foreground="white")
        # cập nhật màu cho các nút đã tạo
        for btn, _, _, _ in self.buttons:
            try:
                btn.configure(activebackground=self.primary_dark)
            except:
                pass
        self._set_status(f"Đã đổi màu chủ đạo: {self.primary}")

    # ---------- HEADER ----------
    def _build_header(self):
        h = ttk.Frame(self.root); h.pack(fill="x", padx=12, pady=(12,6))
        try:
            path = resource_path("logo_truong.png")
            if os.path.exists(path):
                if _HAS_PIL:
                    img = Image.open(path).convert("RGBA").resize((56,56), Image.LANCZOS)
                    self.logo_img = ImageTk.PhotoImage(img)
                else:
                    self.logo_img = tk.PhotoImage(file=path)
                self.logo_lbl = tk.Label(h, image=self.logo_img, bg=self.bg_light)
                self.logo_lbl.pack(side="left", padx=(0,10))
        except Exception:
            pass
        ttk.Label(h, text="Quản lý điểm học sinh", style="Title.TLabel").pack(side="left")

        # đổi màu chủ đạo nhanh
        tk.Button(h, text="🎨 Đổi màu", command=self._color_picker, bg="#455a64", fg="white",
                  activebackground="#37474f", padx=10, pady=5).pack(side="right", padx=(8,0))

    # ---------- KPI STRIP ----------
    def _build_kpis(self):
        k = ttk.Frame(self.root); k.pack(fill="x", padx=12, pady=(0,6))

        def card(parent, title, var):
            f = ttk.Frame(parent, padding=8); f.pack(side="left", padx=6)
            ttk.Label(f, text=title, style="KPI.TLabel").pack(anchor="w")
            ttk.Label(f, textvariable=var, style="KPINum.TLabel").pack(anchor="w")
            return f

        card(k, "Tổng số học sinh", self.kpi_total)
        card(k, "Số Học sinh hiển thị", self.kpi_visible)
        card(k, "Giỏi", self.kpi_gioi)
        card(k, "Khá", self.kpi_kha)
        card(k, "Trung bình", self.kpi_tb)
        card(k, "Yếu", self.kpi_yeu)

    # ---------- FORM ----------
    def _build_form(self):
        frm = ttk.LabelFrame(self.root, text="Thông tin học sinh", padding=10)
        frm.pack(fill="x", padx=12, pady=6)

        ttk.Label(frm, text="Họ tên:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        self.ent_name = ttk.Entry(frm, width=28); self.ent_name.grid(row=0, column=1, padx=6, pady=4)

        ttk.Label(frm, text="Lớp:").grid(row=0, column=2, sticky="w", padx=6, pady=4)
        self.ent_class = ttk.Entry(frm, width=12); self.ent_class.grid(row=0, column=3, padx=6, pady=4)

        self.ent_scores = {}
        for i, subj in enumerate(SUBJECTS):
            r = 1 + (i // 3)
            c = (i % 3) * 2
            ttk.Label(frm, text=f"{subj.capitalize()}:").grid(row=r, column=c, sticky="w", padx=6, pady=4)
            # Giữ Entry để tương thích (nền cũ), nhưng bật validate điểm
            e = ttk.Entry(frm, width=6); e.grid(row=r, column=c+1, padx=6, pady=4)
            e.configure(validate="key", validatecommand=(frm.register(self._validate_score), "%P"))
            self.ent_scores[subj] = e

        for col in range(6): frm.grid_columnconfigure(col, weight=1)

    def _validate_score(self, s):
        if s.strip()=="":
            return True
        s = s.replace(",", ".")
        try:
            v = float(s)
            return 0.0 <= v <= 10.0
        except:
            return False

    # ---------- ICONS (tùy chọn) ----------
    def _load_icon(self, name):
        p = os.path.join("icons", name)
        try:
            if os.path.exists(p): return tk.PhotoImage(file=p)
        except Exception:
            return None
        return None

    # ---------- BUTTONS ----------

    def _build_buttons(self):
        # Khung có thanh cuộn ngang
        outer = ttk.LabelFrame(self.root, text="Chức năng", padding=10)
        outer.pack(fill="x", padx=12, pady=6)

        # Canvas + Horizontal Scrollbar
        self.btn_canvas = tk.Canvas(outer, height=56, highlightthickness=0, bg=self.bg_light)
        self.btn_hsb = ttk.Scrollbar(outer, orient="horizontal", command=self.btn_canvas.xview)
        self.btn_canvas.configure(xscrollcommand=self.btn_hsb.set)

        self.btn_canvas.pack(fill="x", expand=False, side="top")
        self.btn_hsb.pack(fill="x", side="bottom")

        # Frame chứa nút, nằm trong Canvas
        self.btn_frame = tk.Frame(self.btn_canvas, bg=self.bg_light)
        self.btn_window = self.btn_canvas.create_window((0, 0), window=self.btn_frame, anchor="nw")

        def _on_frame_configure(_evt=None):
            # Cập nhật vùng scroll khi nội dung đổi kích thước
            self.btn_canvas.configure(scrollregion=self.btn_canvas.bbox("all"))
        self.btn_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_resize(_evt=None):
            # Giữ khung nút cao và đủ rộng để không dồn dòng
            self.btn_canvas.itemconfig(self.btn_window, height=self.btn_frame.winfo_height())
        self.btn_canvas.bind("<Configure>", _on_canvas_resize)

        # Kéo ngang bằng Shift + con lăn chuột
        def _on_shift_wheel(evt):
            try:
                delta = evt.delta
                # Windows/Mac: evt.delta = ±120 multiples
                step = -1 if delta > 0 else 1
            except Exception:
                step = 1
            self.btn_canvas.xview_scroll(step, "units")
        self.btn_canvas.bind_all("<Shift-MouseWheel>", _on_shift_wheel)

        # Nếu muốn kéo bằng chuột giữa (nhấn giữ) cũng có thể thêm sau

        # Tải icon (nếu có)
        ic_add    = self._load_icon("add.png")
        ic_edit   = self._load_icon("edit.png")
        ic_delete = self._load_icon("delete.png")
        ic_save   = self._load_icon("save.png")
        ic_open   = self._load_icon("open.png")
        ic_excel  = self._load_icon("excel.png")
        ic_clear  = self._load_icon("clear.png")
        ic_chart  = self._load_icon("chart.png")
        ic_cols   = self._load_icon("columns.png")

        def mkbtn(text, cmd, bg, icon=None, fallback=""):
            label = f" {text}" if icon else f"{fallback} {text}"
            b = tk.Button(self.btn_frame, text=label, image=icon, compound="left",
                command=cmd, bg=bg, fg="white",
                activebackground=self.primary_dark,
                font=("Segoe UI", 10, "bold"), padx=10, pady=5)
            b.pack(side="left", padx=6)
            self.buttons.append((b, bg, "white", self.primary_dark))
            return b

        mkbtn("Thêm (Ctrl+N)", self.add_student, self.primary, ic_add, "➕")
        mkbtn("Sửa (Ctrl+E)", self.edit_student, self.primary, ic_edit, "✏️")
        mkbtn("Xóa (Del)", self.delete_student, self.primary, ic_delete, "🗑️")
        mkbtn("Lưu (Ctrl+S)", self.save_csv, "#1565c0", ic_save, "💾")
        mkbtn("Đọc (Ctrl+O)", self.load_csv_or_xlsx, "#1565c0", ic_open, "📂")
        mkbtn("Xuất File", lambda: self.export_excel(subset_only=False), "#2e7d32", ic_excel, "📊")
        mkbtn("Xuất File đang hiển thị", lambda: self.export_excel(subset_only=True), "#388e3c", ic_excel, "📑")
        mkbtn("Biểu đồ", self.show_charts, "#425862", ic_chart, "📈")
        mkbtn("Biểu đồ 3 khối", self.show_block_chart, "#455a64", ic_chart, "🏫")
        mkbtn("Ẩn/Hiện cột", self.toggle_columns_dialog, "#546e7a", ic_cols, "🧩")
        mkbtn("Hỏi AI (Ctrl+Q)", lambda: open_qna_window(self.root, get_subset_callable=self._get_visible_subset), "#6a1b9a", None, "🤖")
        mkbtn("Dark Mode (Ctrl+D)", self.toggle_dark_mode, self.primary, None, "🌙")

        b_clear = tk.Button(self.btn_frame, text=(" Xóa form" if ic_clear else "🧹 Xóa form"),
                            image=ic_clear, compound="left",
                            command=self.clear_form, bg="#90caf9", fg="#0b3060",
                            activebackground="#64b5f6",
                            font=("Segoe UI", 10, "bold"), padx=10, pady=5)
        b_clear.pack(side="left", padx=16)
        self.buttons.append((b_clear, "#90caf9", "#0b3060", "#64b5f6"))

    # ---------- SEARCH ----------
    def _build_search(self):
        sf = ttk.LabelFrame(self.root, text="Tìm kiếm", padding=10)
        sf.pack(fill="x", padx=12, pady=6)

        ttk.Label(sf, text="Tên:").pack(side="left", padx=4)
        self.ent_search_name = ttk.Entry(sf, width=14); self.ent_search_name.pack(side="left")

        ttk.Label(sf, text="Lớp:").pack(side="left", padx=4)
        self.ent_search_class = ttk.Entry(sf, width=10); self.ent_search_class.pack(side="left")

        ttk.Label(sf, text="Xếp loại:").pack(side="left", padx=4)
        self.ent_search_rank = ttk.Entry(sf, width=12); self.ent_search_rank.pack(side="left")

        b_adv = tk.Button(sf, text="🔎 Tìm nâng cao",
                          command=self.advanced_search, bg="#00897b", fg="white",
                          activebackground="#00695c",
                          font=("Segoe UI", 10, "bold"), padx=10, pady=4)
        b_adv.pack(side="left", padx=8)
        self.buttons.append((b_adv, "#00897b", "white", "#00695c"))

        ttk.Label(sf, text="Theo:").pack(side="left", padx=(16,0))
        self.cmb_criteria = ttk.Combobox(
            sf, state="readonly", width=12,
            values=["Tên", "Lớp", "ID", "Xếp loại"]
        )
        self.cmb_criteria.current(0); self.cmb_criteria.pack(side="left", padx=6)
        ttk.Label(sf, text="Giá trị:").pack(side="left", padx=(10,0))
        self.ent_search = ttk.Entry(sf, width=26); self.ent_search.pack(side="left", padx=6)

        b_find = tk.Button(sf, text="🔍 Tìm (Ctrl+F)",
                           command=self.search_student, bg=self.primary, fg="white",
                           activebackground=self.primary_dark,
                           font=("Segoe UI", 10, "bold"), padx=10, pady=4)
        b_find.pack(side="left", padx=6)
        self.buttons.append((b_find, self.primary, "white", self.primary_dark))

        b_all = tk.Button(sf, text="📋 Hiển thị tất cả (F5)",
                          command=self.refresh_table, bg=self.primary, fg="white",
                          activebackground=self.primary_dark,
                          font=("Segoe UI", 10, "bold"), padx=10, pady=4)
        b_all.pack(side="left", padx=6)
        self.buttons.append((b_all, self.primary, "white", self.primary_dark))

    # ---------- TABLE ----------
    def _build_table(self):
        wrap = ttk.Frame(self.root); wrap.pack(fill="both", expand=True, padx=12, pady=6)
        self.cols = ["stt","id","ho_ten","lop"] + SUBJECTS + ["diem_tb","xep_loai"]
        self.tree = ttk.Treeview(wrap, columns=self.cols, show="headings", height=18)
        self.tree["displaycolumns"] = self.display_columns

        header_vi = {
            "stt": "STT", "id": "ID", "ho_ten": "Họ Tên", "lop": "Lớp",
            "toan": "Toán", "ly": "Lý", "hoa": "Hóa", "van": "Văn",
            "anh": "Anh", "tin": "Tin", "diem_tb": "Điểm TB", "xep_loai": "Xếp Loại"
        }
        for key in self.cols:
            self.tree.heading(key, text=header_vi.get(key, key),
                              command=lambda k=key: self._on_sort_column(k))

        self.tree.column("stt", width=56, anchor="center")
        self.tree.column("id", width=70, anchor="center")
        self.tree.column("ho_ten", width=240, anchor="w")
        self.tree.column("lop", width=110, anchor="center")
        for s in SUBJECTS: self.tree.column(s, width=85, anchor="e")
        self.tree.column("diem_tb", width=95, anchor="e")
        self.tree.column("xep_loai", width=110, anchor="center")

        vsb = ttk.Scrollbar(wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1); wrap.columnconfigure(0, weight=1)

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Button-3>", self._on_right_click)

        self._configure_row_tags(dark=False)

    # ---------- STATUS ----------
    def _build_statusbar(self):
        self.status = tk.StringVar()
        bar = ttk.Frame(self.root); bar.pack(fill="x", side="bottom")
        ttk.Label(bar, textvariable=self.status, anchor="w").pack(fill="x", padx=12, pady=4)
    def _set_status(self, msg): self.status.set(msg)

    # ---------- SHORTCUTS ----------
    def _bind_shortcuts(self):
        self.root.bind("<Control-n>", lambda e: self.add_student())
        self.root.bind("<Control-N>", lambda e: self.add_student())
        self.root.bind("<Control-e>", lambda e: self.edit_student())
        self.root.bind("<Control-E>", lambda e: self.edit_student())
        self.root.bind("<Delete>",    lambda e: self.delete_student())
        self.root.bind("<Control-s>", lambda e: self.save_csv())
        self.root.bind("<Control-S>", lambda e: self.save_csv())
        self.root.bind("<Control-o>", lambda e: self.load_csv_or_xlsx())
        self.root.bind("<Control-O>", lambda e: self.load_csv_or_xlsx())
        self.root.bind("<Control-d>", self.toggle_dark_mode)
        self.root.bind("<Control-D>", self.toggle_dark_mode)
        self.root.bind("<Control-f>", lambda e: (self.ent_search.focus_set(), "break"))
        self.root.bind("<F5>",        lambda e: self.refresh_table())
        self.root.bind("<Control-q>", lambda e: open_qna_window(self.root, get_subset_callable=self._get_visible_subset))

    # ---------- HELPERS ----------
    def clear_form(self):
        self.ent_name.delete(0, tk.END); self.ent_class.delete(0, tk.END)
        for e in self.ent_scores.values(): e.delete(0, tk.END)
        self._set_status("Đã xóa nội dung form.")

    def _on_tree_select(self, _evt):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])["values"]
        self.ent_name.delete(0, tk.END); self.ent_name.insert(0, vals[2])
        self.ent_class.delete(0, tk.END); self.ent_class.insert(0, vals[3])
        for i, subj in enumerate(SUBJECTS, start=4):
            self.ent_scores[subj].delete(0, tk.END)
            self.ent_scores[subj].insert(0, vals[i])
        self._set_status(f"Đang chọn ID {vals[1]}.")

    def _collect_scores(self):
        return {s: parse_score_any(e.get()) for s, e in self.ent_scores.items()}

    def _on_sort_column(self, col):
        # Bỏ qua cột STT vì là thứ tự hiển thị
        if col == "stt":
            return
        ascending = self.sort_state.get(col, True)
        key_fn = None
        if col in SUBJECTS + ["diem_tb"]:
            key_fn = lambda s: float(s.get(col, 0.0))
        elif col in ["id"]:
            key_fn = lambda s: int(s.get(col, 0))
        else:
            key_fn = lambda s: str(s.get(col, "")).lower()
        self.students.sort(key=key_fn, reverse=not ascending)
        self.sort_state[col] = not ascending
        self.refresh_table()
        self._set_status(f"Sắp xếp theo '{col}' ({'↑' if ascending else '↓'})")

    def _on_right_click(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="✏️ Sửa", command=self.edit_student)
        menu.add_command(label="🗑️ Xóa", command=self.delete_student)
        menu.add_separator()
        menu.add_command(label="📋 Copy hàng", command=self._copy_selected_row)
        menu.add_command(label="📋 Copy ô", command=lambda: self._copy_cell(event))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _copy_selected_row(self):
        sel = self.tree.selection()
        if not sel: return
        vals = self.tree.item(sel[0])["values"]
        txt = "; ".join(str(v) for v in vals)
        self.root.clipboard_clear()
        self.root.clipboard_append(txt)
        self._set_status("Đã copy hàng.")

    def _copy_cell(self, event):
        col = self.tree.identify_column(event.x)  # e.g. '#3'
        row = self.tree.identify_row(event.y)
        if not col or not row: return
        idx = int(col.replace("#","")) - 1
        vals = self.tree.item(row)["values"]
        if idx < 0 or idx >= len(vals): return
        self.root.clipboard_clear()
        self.root.clipboard_append(str(vals[idx]))
        self._set_status("Đã copy ô.")

    # ---------- CRUD ----------
    def add_student(self):
        name = self.ent_name.get().strip(); lop = self.ent_class.get().strip()
        if not name or not lop:
            messagebox.showwarning("Thiếu", "Vui lòng nhập Họ tên và Lớp."); return
        scores = self._collect_scores()
        avg = wavg(scores); xl = classify(avg)
        st = {"id": self.next_id, "ho_ten": name, "lop": lop,
              **scores, "diem_tb": round(avg, 2), "xep_loai": xl}
        self.students.append(st); self.next_id += 1
        self.refresh_table(); self._set_status(f"Đã thêm HS {name} ({lop}).")

    def edit_student(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Chọn dòng", "Chọn một học sinh trong bảng để sửa."); return
        sid = int(self.tree.item(sel[0])["values"][1])
        st = next((s for s in self.students if s["id"] == sid), None)
        if not st: return
        name = self.ent_name.get().strip() or st["ho_ten"]
        lop  = self.ent_class.get().strip() or st["lop"]
        new_scores = self._collect_scores()
        for k in SUBJECTS:
            if self.ent_scores[k].get().strip() == "":
                new_scores[k] = st[k]
        avg = wavg(new_scores); xl = classify(avg)
        st.update({"ho_ten": name, "lop": lop, **new_scores,
                   "diem_tb": round(avg, 2), "xep_loai": xl})
        self.refresh_table(); self._set_status(f"Đã sửa ID {sid}.")

    def delete_student(self):
        sel = self.tree.selection()
        if not sel: return
        sid = int(self.tree.item(sel[0])["values"][1])
        self.students = [s for s in self.students if s["id"] != sid]
        self.refresh_table(); self._set_status(f"Đã xóa ID {sid}.")

    # ---------- REFRESH TABLE + KPI + HISTOGRAM ----------
    def refresh_table(self, subset=None):
        for i in self.tree.get_children(): self.tree.delete(i)

        global TOAN_RAW, LI_RAW, HOA_RAW, VAN_RAW, ANH_RAW, TIN_RAW, TB_RAW
        TOAN_RAW, LI_RAW, HOA_RAW = [], [], []
        VAN_RAW, ANH_RAW, TIN_RAW = [], [], []
        TB_RAW = []

        data = subset if subset is not None else self.students
        data.sort(key=lambda s: (s["lop"], s["ho_ten"]))

        rank_count = {"Giỏi":0,"Khá":0,"Trung bình":0,"Yếu":0}

        for idx, s in enumerate(data, start=1):
            row = [idx, s["id"], s["ho_ten"], s["lop"]] + \
                  [s[subj] for subj in SUBJECTS] + [f"{s['diem_tb']:.2f}", s["xep_loai"]]

            rank_tag = {
                "Giỏi": "rank_gioi",
                "Khá": "rank_kha",
                "Trung bình": "rank_tb",
                "Yếu": "rank_yeu"
            }.get(s["xep_loai"], "oddrow")
            rank_count[s["xep_loai"]] = rank_count.get(s["xep_loai"],0) + 1

            zebra = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=row, tags=(zebra, rank_tag))

            TOAN_RAW.append(s["toan"]); LI_RAW.append(s["ly"]); HOA_RAW.append(s["hoa"])
            VAN_RAW.append(s["van"]); ANH_RAW.append(s["anh"]); TIN_RAW.append(s["tin"])
            TB_RAW.append(s["diem_tb"])

        update_histograms_from_raw()

        # KPI update
        self.kpi_total.set(str(len(self.students)))
        self.kpi_visible.set(str(len(data)))
        self.kpi_gioi.set(str(rank_count.get("Giỏi",0)))
        self.kpi_kha.set(str(rank_count.get("Khá",0)))
        self.kpi_tb.set(str(rank_count.get("Trung bình",0)))
        self.kpi_yeu.set(str(rank_count.get("Yếu",0)))

        # theme row tags
        self._configure_row_tags(self.dark_mode)

    # ---------- TÌM KIẾM THƯỜNG ----------
    def search_student(self):
        crit = self.cmb_criteria.get(); q = self.ent_search.get().strip()
        if not q:
            self.refresh_table(); self._set_status("Hiển thị tất cả."); return

        if crit == "Tên":
            ql = q.lower(); filtered = [s for s in self.students if ql in s["ho_ten"].lower()]
        elif crit == "Lớp":
            ql = q.lower(); filtered = [s for s in self.students if ql in s["lop"].lower()]
        elif crit == "ID":
            try:
                qid = int(q); filtered = [s for s in self.students if s["id"] == qid]
            except ValueError:
                messagebox.showwarning("ID không hợp lệ", "Nhập số nguyên cho ID."); return
        elif crit == "Xếp loại":
            ql = q.lower(); filtered = [s for s in self.students if ql in s["xep_loai"].lower()]
        else:
            filtered = self.students

        self.refresh_table(filtered)
        self._set_status(f"Tìm theo {crit}='{q}' → {len(filtered)} kết quả.")

    # ---------- TÌM KIẾM NÂNG CAO ----------
    def advanced_search(self):
        qname  = self.ent_search_name.get().strip().lower()
        qclass = self.ent_search_class.get().strip().lower()
        qrank  = self.ent_search_rank.get().strip().lower()

        filtered = []
        for s in self.students:
            ok = True
            if qname  and qname  not in s["ho_ten"].lower(): ok = False
            if qclass and qclass not in s["lop"].lower():    ok = False
            if qrank  and qrank  not in s["xep_loai"].lower(): ok = False
            if ok: filtered.append(s)

        self.refresh_table(filtered)
        self._set_status(f"Tìm nâng cao → {len(filtered)} kết quả.")

    # ---------- CSV / XLSX ----------
    def save_csv(self):
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                            filetypes=[("CSV files","*.csv"), ("All files","*.*")])
        if not path: return
        header_vi = {
            "id": "ID", "ho_ten": "Họ Tên", "lop": "Lớp",
            "toan": "Toán", "ly": "Lý", "hoa": "Hóa", "van": "Văn",
            "anh": "Anh", "tin": "Tin", "diem_tb": "Điểm TB", "xep_loai": "Xếp Loại"
        }
        cols_en = ["id","ho_ten","lop"] + SUBJECTS + ["diem_tb","xep_loai"]
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
                w.writerow([header_vi[c] for c in cols_en])
                for s in self.students: w.writerow([s[c] for c in cols_en])
            messagebox.showinfo("Thành công", f"Đã lưu {len(self.students)} HS vào:\n{path}")
            self._set_status(f"Đã lưu CSV: {path}")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def load_csv_or_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("CSV/XLSX","*.csv *.xlsx"), ("CSV","*.csv"), ("Excel","*.xlsx"), ("All files","*.*")])
        if not path: return
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".xlsx":
                self._load_xlsx(path)
            else:
                self._load_csv(path)
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def _load_csv(self, path):
        self.students.clear()
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096); f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,")
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ";"
            reader = csv.DictReader(f, delimiter=delimiter, quotechar='"')
            vi_to_en = {
                "id":"id","họ tên":"ho_ten","ho ten":"ho_ten",
                "lớp":"lop","lop":"lop","toán":"toan","toan":"toan",
                "lý":"ly","ly":"ly","hóa":"hoa","hoa":"hoa",
                "văn":"van","van":"van","anh":"anh","tin":"tin",
                "điểm tb":"diem_tb","diem tb":"diem_tb",
                "xếp loại":"xep_loai","xep loai":"xep_loai"
            }
            file_fields = reader.fieldnames or []
            field_map = {}
            for h in file_fields:
                k = (h or "").strip().lower()
                en = vi_to_en.get(k)
                if not en and k in ["id","ho_ten","lop","toan","ly","hoa","van","anh","tin","diem_tb","xep_loai"]:
                    en = k
                if en: field_map[h] = en

            for row in reader:
                get = lambda en_key: next((row[h] for h, en in field_map.items() if en == en_key), "")
                sc = {s: parse_score_any(get(s)) for s in SUBJECTS}
                avg = wavg(sc); xl = classify(avg)
                id_val = get("id")
                try: id_int = int(float(id_val)) if id_val != "" else len(self.students)+1
                except: id_int = len(self.students)+1
                self.students.append({
                    "id": id_int,
                    "ho_ten": (get("ho_ten") or "").strip(),
                    "lop": (get("lop") or "").strip(),
                    **sc, "diem_tb": round(avg, 2), "xep_loai": xl
                })
        self.next_id = max((s["id"] for s in self.students), default=0) + 1
        self.refresh_table()
        self._set_status(f"Đã đọc CSV: {path}")

    def _load_xlsx(self, path):
        self.students.clear()
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Lấy header
        headers = []
        for c in ws[1]:
            headers.append((c.value or "").strip() if isinstance(c.value, str) else str(c.value or ""))
        vi_to_en = {
            "id":"id","họ tên":"ho_ten","ho ten":"ho_ten",
            "lớp":"lop","lop":"lop","toán":"toan","toan":"toan",
            "lý":"ly","ly":"ly","hóa":"hoa","hoa":"hoa",
            "văn":"van","van":"van","anh":"anh","tin":"tin",
            "điểm tb":"diem_tb","diem tb":"diem_tb",
            "xếp loại":"xep_loai","xep loai":"xep_loai"
        }
        # map cột
        col_map = {}
        for idx, h in enumerate(headers):
            k = h.strip().lower()
            en = vi_to_en.get(k) or (k if k in ["id","ho_ten","lop","toan","ly","hoa","van","anh","tin","diem_tb","xep_loai"] else None)
            if en:
                col_map[idx] = en

        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {col_map[i]: r[i] for i in col_map.keys() if i < len(r)}
            get = lambda en_key, default="": row.get(en_key, default)
            sc = {s: parse_score_any(get(s, 0)) for s in SUBJECTS}
            avg = wavg(sc); xl = classify(avg)
            id_val = get("id")
            try: id_int = int(float(id_val)) if id_val not in ("", None) else len(self.students)+1
            except: id_int = len(self.students)+1
            self.students.append({
                "id": id_int,
                "ho_ten": (str(get("ho_ten") or "")).strip(),
                "lop": (str(get("lop") or "")).strip(),
                **sc, "diem_tb": round(avg, 2), "xep_loai": xl
            })
        self.next_id = max((s["id"] for s in self.students), default=0) + 1
        self.refresh_table()
        self._set_status(f"Đã đọc Excel: {path}")

    # ---------- EXCEL ----------
    def export_excel(self, subset_only=False):
        data = self._get_visible_subset() if subset_only else self.students
        if not data:
            messagebox.showinfo("Trống", "Chưa có dữ liệu để xuất."); return

        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel Workbook","*.xlsx")])
        if not path: return

        headers = ["STT","ID","Họ Tên","Lớp","Toán","Lý","Hóa","Văn","Anh","Tin","Điểm TB","Xếp Loại"]
        wb = Workbook(); ws = wb.active; ws.title = "Bảng điểm"

        head_font = Font(bold=True)
        head_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        zebra_even = PatternFill(start_color="F7FBFF", end_color="F7FBFF", fill_type="solid")
        zebra_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        thin = Side(style="thin", color="90A4AE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = head_font; cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        data_sorted = sorted(data, key=lambda s: (s["lop"], s["ho_ten"]))
        for idx, s in enumerate(data_sorted, start=1):
            row_vals = [idx, s["id"], s["ho_ten"], s["lop"],
                        s["toan"], s["ly"], s["hoa"], s["van"], s["anh"], s["tin"],
                        float(f"{s['diem_tb']:.2f}"), s["xep_loai"]]
            r = idx + 1
            for c, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                if c in (1,2,4,11): cell.alignment = Alignment(horizontal="center")
                elif c == 3: cell.alignment = Alignment(horizontal="left")
                else: cell.alignment = Alignment(horizontal="center")
                cell.fill = zebra_even if idx % 2 == 0 else zebra_odd
                cell.border = border

        for column_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 28)

        try:
            wb.save(path)
            messagebox.showinfo("Thành công", f"Đã xuất Excel{' (đang hiển thị)' if subset_only else ''}:\n{path}")
            self._set_status(f"Đã xuất Excel: {path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu Excel:\n{e}")

    # ---------- LẤY TẬP DỮ LIỆU ĐANG HIỂN THỊ ----------
    def _get_visible_subset(self):
        id_to_obj = {s["id"]: s for s in self.students}
        subset = []
        for item in self.tree.get_children():
            vals = self.tree.item(item)["values"]
            if not vals:
                continue
            try:
                sid = int(vals[1])
                if sid in id_to_obj:
                    subset.append(id_to_obj[sid])
            except Exception:
                continue
        return subset

    # ---------- ẨN/HIỆN CỘT ----------
    def toggle_columns_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Ẩn/Hiện cột")
        dlg.geometry("320x360")
        ttk.Label(dlg, text="Chọn cột muốn hiển thị:", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=10, pady=8)
        vars_map = {}
        for c in self.cols:
            var = tk.BooleanVar(value=(c in self.display_columns))
            cb = ttk.Checkbutton(dlg, text=c, variable=var)
            cb.pack(anchor="w", padx=14, pady=2)
            vars_map[c] = var
        def apply():
            shown = [c for c,v in vars_map.items() if v.get()]
            if "stt" not in shown:
                shown.insert(0, "stt")
            self.display_columns = shown
            self.tree["displaycolumns"] = shown
            dlg.destroy()
        ttk.Button(dlg, text="Áp dụng", command=apply).pack(pady=10)

    # ---------- BIỂU ĐỒ HISTOGRAM 7 MÔN ----------
    def show_charts(self):
        subset = self._get_visible_subset()
        if not subset:
            messagebox.showinfo("Chưa có dữ liệu", "Bảng đang rỗng. Hãy thêm học sinh hoặc hiển thị dữ liệu trước.")
            return

        def hist10(values):
            bins = [0]*10
            for v in values:
                bins[_bucket_index(v)] += 1
            return [int(x) for x in bins]

        tb_vals   = [s["diem_tb"] for s in subset]
        toan_vals = [s["toan"]    for s in subset]
        ly_vals   = [s["ly"]      for s in subset]
        hoa_vals  = [s["hoa"]     for s in subset]
        van_vals  = [s["van"]     for s in subset]
        anh_vals  = [s["anh"]     for s in subset]
        tin_vals  = [s["tin"]     for s in subset]

        y_tb   = hist10(tb_vals)
        y_toan = hist10(toan_vals)
        y_li   = hist10(ly_vals)
        y_hoa  = hist10(hoa_vals)
        y_van  = hist10(van_vals)
        y_anh  = hist10(anh_vals)
        y_tin  = hist10(tin_vals)

        x_labels = ["0-1","1-2","2-3","3-4","4-5","5-6","6-7","7-8","8-9","9-10"]
        series = [
            ("TRUNG BÌNH MÔN", y_tb,   "blue"),
            ("TOÁN HỌC",       y_toan, "green"),
            ("VẬT LÝ",         y_li,   "red"),
            ("HÓA HỌC",        y_hoa,  "orange"),
            ("NGỮ VĂN",        y_van,  "purple"),
            ("NGOẠI NGỮ",      y_anh,  "gold"),
            ("TIN HỌC",        y_tin,  "grey"),
        ]

        fig, axs = plt.subplots(4, 2, figsize=(12, 10))
        axs = axs.ravel()

        for i, (title, data, color) in enumerate(series):
            bars = axs[i].bar(x_labels, data, color=color, width=0.6)
            axs[i].set_title(title)
            axs[i].set_ylabel("Số lượng")
            axs[i].set_xlabel("Khoảng điểm")
            axs[i].grid(axis="y", linestyle="--", alpha=0.3)
            axs[i].yaxis.set_major_locator(MaxNLocator(integer=True))
            y_max = max(data) if max(data) > 0 else 1
            axs[i].set_ylim(0, y_max * 1.2 + 1)
            y_off = max(0.4, y_max * 0.03)
            for rect, val in zip(bars, data):
                axs[i].text(rect.get_x() + rect.get_width()/2.0,
                            rect.get_height() + y_off,
                            str(int(val)),
                            ha="center", va="bottom", fontsize=9, clip_on=True)

        axs[-1].axis("off")
        fig.suptitle(f"Phân bố điểm theo khoảng – Số học sinh hiển thị: {len(subset)}", y=0.995)
        plt.tight_layout()
        try:
            fig.savefig("bieudo.png", dpi=300, transparent=True, bbox_inches="tight", pad_inches=0.2)
        except Exception:
            pass
        plt.show()

    # ---------- BIỂU ĐỒ 3 KHỐI ----------
    def show_block_chart(self):
        subset = self._get_visible_subset()
        if not subset:
            messagebox.showinfo("Chưa có dữ liệu", "Bảng đang rỗng. Hãy thêm học sinh hoặc hiển thị dữ liệu trước.")
            return

        # gom theo khối dựa vào tiền tố lớp
        blocks = {"10": [], "11": [], "12": []}
        for s in subset:
            lop = (s.get("lop") or "").strip()
            tb  = float(s.get("diem_tb", 0.0))
            if lop.startswith("10"):
                blocks["10"].append(tb)
            elif lop.startswith("11"):
                blocks["11"].append(tb)
            elif lop.startswith("12"):
                blocks["12"].append(tb)

        # histogram 10 bin
        def hist10(values):
            bins = [0]*10
            for v in values:
                try: x = float(v)
                except: x = 0.0
                if x < 0: x = 0.0
                if x > 10: x = 10.0
                idx = 9 if x >= 9.0 else int(x)
                bins[idx] += 1
            return [int(x) for x in bins]

        h10 = hist10(blocks["10"])
        h11 = hist10(blocks["11"])
        h12 = hist10(blocks["12"])
        havg = [(h10[i] + h11[i] + h12[i]) / 3.0 for i in range(10)]  # TB theo từng bin

        x_labels = ["0-1","1-2","2-3","3-4","4-5","5-6","6-7","7-8","8-9","9-10"]
        x = np.arange(len(x_labels))
        width = 0.2

        fig, ax = plt.subplots(figsize=(12, 6))
        r10  = ax.bar(x - 1.5*width, h10, width, label="Khối 10")
        r11  = ax.bar(x - 0.5*width, h11, width, label="Khối 11")
        r12  = ax.bar(x + 0.5*width, h12, width, label="Khối 12")
        ravg = ax.bar(x + 1.5*width, havg, width, label="TB 3 khối")

        ax.set_title("Phân bố Điểm TB theo khoảng – 3 khối & Trung bình cộng")
        ax.set_xlabel("Khoảng điểm")
        ax.set_ylabel("Số học sinh")
        ax.set_xticks(x, x_labels)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.grid(axis="y", linestyle="--", alpha=0.3)
        ax.legend(loc="upper right", ncols=2)

        ymax = max(max(h10 or [0]), max(h11 or [0]), max(h12 or [0]), int(max(havg or [0])))
        ymax = 1 if ymax <= 0 else ymax
        ax.set_ylim(0, ymax*1.25 + 1)

        def annotate(bars):
            for rect in bars:
                h = rect.get_height()
                ax.text(rect.get_x() + rect.get_width()/2,
                        h + max(0.4, ymax*0.03),
                        f"{h:.0f}" if abs(h - round(h)) < 1e-9 else f"{h:.1f}",
                        ha="center", va="bottom", fontsize=9, clip_on=True)
        annotate(r10); annotate(r11); annotate(r12); annotate(ravg)

        plt.tight_layout()
        plt.show()

# ---------- AI Q&A WINDOW (giữ nguyên từ nền cũ để tương thích) ----------
def open_qna_window(root, get_subset_callable=None):
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    from tkinter.scrolledtext import ScrolledText
    from datetime import datetime
    import os

    win = tk.Toplevel(root)
    win.title("Hỏi AI (Q&A)")
    win.geometry("1100x740")

    style = ttk.Style(win)
    style.configure("QnA.TLabelframe", padding=10)
    style.configure("QnA.TLabelframe.Label", font=("Segoe UI", 11, "bold"))
    style.configure("QnA.Primary.TButton",  foreground="white", background="#1e88e5")
    style.map("QnA.Primary.TButton",
              background=[("active", "#1565c0"), ("!active", "#1e88e5")])
    style.configure("QnA.Info.TButton",     foreground="white", background="#0288d1")
    style.map("QnA.Info.TButton",
              background=[("active", "#0277bd"), ("!active", "#0288d1")])
    style.configure("QnA.Secondary.TButton", foreground="white", background="#607d8b")
    style.map("QnA.Secondary.TButton",
              background=[("active", "#546e7a"), ("!active", "#607d8b")])

    log_lines = []
    history_rows = []  # (time, question, answer, idx)

    def log(msg: str):
        from datetime import datetime
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        log_lines.append(line)
        txt_log.configure(state="normal")
        txt_log.insert("end", line + "\n")
        txt_log.see("end")
        txt_log.configure(state="disabled")

    nb = ttk.Notebook(win)
    tab_chat = ttk.Labelframe(nb, text="💬 Chat với AI", style="QnA.TLabelframe")
    tab_ctx  = ttk.Labelframe(nb, text="📄 Ngữ cảnh",   style="QnA.TLabelframe")
    tab_cfg  = ttk.Labelframe(nb, text="⚙️ Cài đặt",    style="QnA.TLabelframe")
    tab_log  = ttk.Labelframe(nb, text="🧾 Nhật ký",    style="QnA.TLabelframe")
    nb.add(tab_chat, text="Chat"); nb.add(tab_ctx, text="Ngữ cảnh")
    nb.add(tab_cfg, text="Cài đặt"); nb.add(tab_log, text="Nhật ký")
    nb.pack(fill="both", expand=True, padx=10, pady=10)

    left = ttk.Labelframe(tab_chat, text="Ngân hàng câu hỏi", style="QnA.TLabelframe")
    left.pack(side="left", fill="y", padx=(0,10))
    questions = [
        "Phân tích phân bố xếp loại theo lớp.",
        "Liệt kê top 10 học sinh theo Điểm TB.",
        "Tìm các lớp có Điểm TB trung bình ≥ 8.0.",
        "Nhận xét chênh lệch điểm Toán-Văn.",
        "Lọc HS khối 11 có Điểm TB ≥ 7.5.",
        "Mô tả báo cáo theo khối 10/11/12.",
        "Tìm HS điểm thấp nhất từng môn.",
        "Gợi ý biểu đồ phù hợp dữ liệu này.",
        "Tạo tiêu đề & tóm tắt báo cáo tuần.",
        "Đề xuất tiêu chí xét học bổng."
    ]
    lst = tk.Listbox(left, height=22)
    for q in questions: lst.insert("end", q)
    lst.pack(fill="y", expand=True)

    right = ttk.Frame(tab_chat)
    right.pack(side="left", fill="both", expand=True)

    frm_q = ttk.Labelframe(right, text="Câu hỏi", style="QnA.TLabelframe")
    frm_q.pack(fill="x")
    txt_q = ScrolledText(frm_q, height=5, wrap="word", font=("Segoe UI", 11), relief="flat")
    txt_q.pack(fill="x", pady=(0,6))

    def insert_from_bank(_evt=None):
        sel = lst.curselection()
        if not sel: return
        q = lst.get(sel[0])
        txt_q.insert("end", ("" if txt_q.index("end-1c")=="1.0" else "\n") + q)
        txt_q.focus_set()
    lst.bind("<Double-Button-1>", insert_from_bank)

    ttk.Separator(right, orient="horizontal").pack(fill="x", pady=6)

    frm_a = ttk.Labelframe(right, text="Trả lời", style="QnA.TLabelframe")
    frm_a.pack(fill="both", expand=True)
    txt_a = ScrolledText(frm_a, height=14, wrap="word", font=("Segoe UI", 11), relief="flat")
    txt_a.pack(fill="both", expand=True)

    action = ttk.Frame(right); action.pack(fill="x", pady=(8,0))
    include_ctx_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(action, text="Đưa ngữ cảnh bảng vào câu hỏi",
                    variable=include_ctx_var).pack(side="left")

    def on_copy():
        win.clipboard_clear()
        win.clipboard_append(txt_a.get("1.0","end").strip())
        log("Đã copy trả lời.")

    def on_export_log():
        path = filedialog.asksaveasfilename(defaultextension=".txt",
                                            filetypes=[("Text","*.txt")])
        if not path: return
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(log_lines))
        messagebox.showinfo("Xuất", f"Đã lưu log:\n{path}")

    btn_ask    = ttk.Button(action, text="🤖 Hỏi AI (Ctrl+Enter)")
    btn_copy   = ttk.Button(action, text="📋 Copy",                 command=on_copy)
    btn_export = ttk.Button(action, text="💾 Xuất log",             command=on_export_log)
    btn_export.pack(side="right", padx=6); btn_copy.pack(side="right", padx=6); btn_ask.pack(side="right", padx=6)

    history = ttk.Labelframe(tab_chat, text="Lịch sử hỏi đáp", style="QnA.TLabelframe")
    history.pack(side="left", fill="y", padx=(10,0))
    tv = ttk.Treeview(history, columns=("time","q","a","idx"), show="headings", height=25)
    tv.heading("time", text="Thời gian"); tv.heading("q", text="Câu hỏi"); tv.heading("a", text="Tóm tắt trả lời")
    tv.column("time", width=85, anchor="center")
    tv.column("q", width=280, anchor="w"); tv.column("a", width=280, anchor="w")
    tv.heading("idx", text=""); tv.column("idx", width=0, stretch=False)
    tv.pack(fill="y", expand=True)

    def on_select_history(_evt=None):
        sel = tv.selection()
        if not sel: return
        vals = tv.item(sel[0])["values"]
        if len(vals) < 4: return
        idx = int(vals[3])
        t,q,a,_ = history_rows[idx]
        txt_q.delete("1.0","end"); txt_q.insert("end", q)
        txt_a.delete("1.0","end"); txt_a.insert("end", a)
    tv.bind("<<TreeviewSelect>>", on_select_history)

    txt_ctx = ScrolledText(tab_ctx, wrap="none", font=("Consolas", 10))
    txt_ctx.pack(fill="both", expand=True)
    def refresh_context():
        txt_ctx.delete("1.0","end")
        if callable(get_subset_callable):
            try:
                subset = get_subset_callable() or []
                txt_ctx.insert("end", _ai__summarize_records(subset, limit_rows=50))
            except Exception as e:
                txt_ctx.insert("end", f"(Lỗi lấy ngữ cảnh) {e}")
        else:
            txt_ctx.insert("end", "(Không có hàm truy cập ngữ cảnh)")
    ttk.Button(tab_ctx, text="🔄 Làm mới", command=refresh_context)\
        .pack(anchor="e", pady=6)
    refresh_context()

    cfg = ttk.Frame(tab_cfg, padding=4); cfg.pack(fill="x")
    ttk.Label(cfg, text="Model (ENV OPENAI_MODEL):").grid(row=0, column=0, sticky="w", pady=6)
    model_var = tk.StringVar(value=os.getenv("OPENAI_MODEL", "gpt-4o-mini"))
    model_box = ttk.Combobox(cfg, textvariable=model_var,
                             values=("gpt-4o-mini","gpt-4o","gpt-5-mini"),
                             width=20, state="readonly")
    model_box.grid(row=0, column=1, sticky="w", padx=8)
    def on_model_sel(*_):
        os.environ["OPENAI_MODEL"] = model_var.get()
        log(f"Chọn model: {model_var.get()}")
    model_box.bind("<<ComboboxSelected>>", on_model_sel)

    ttk.Label(cfg, text="Temperature (0.0 - 2.0):").grid(row=1, column=0, sticky="w", pady=6)
    spn_temp = ttk.Spinbox(cfg, from_=0.0, to=2.0, increment=0.1, width=6)
    spn_temp.insert(0, "0.2"); spn_temp.grid(row=1, column=1, sticky="w", padx=8)

    ttk.Label(cfg, text="Max output tokens:").grid(row=2, column=0, sticky="w", pady=6)
    spn_tokens = ttk.Spinbox(cfg, from_=100, to=4000, increment=50, width=8)
    spn_tokens.insert(0, "400"); spn_tokens.grid(row=2, column=1, sticky="w", padx=8)

    txt_log = ScrolledText(tab_log, wrap="word", state="disabled", font=("Segoe UI", 10))
    txt_log.pack(fill="both", expand=True)

    def _shorten(s, n=80):
        s = " ".join(s.split())
        return s if len(s) <= n else s[:n-1] + "…"

    def on_ask(_evt=None):
        q = txt_q.get("1.0","end").strip()
        if not q:
            messagebox.showinfo("Thiếu nội dung", "Hãy nhập câu hỏi."); return

        txt_a.delete("1.0","end"); txt_a.insert("end", "⏳ Đang hỏi AI...\n")

        ctx = ""
        if callable(get_subset_callable):
            try: ctx = _ai__summarize_records(get_subset_callable() or [], limit_rows=15)
            except: ctx = ""

        try:
            temp = float(spn_temp.get()); toks = int(spn_tokens.get())
        except: temp, toks = 0.2, 400

        ans = _ai__ask_chatgpt(q, context=ctx, temperature=temp, max_output_tokens=toks)
        txt_a.delete("1.0","end"); txt_a.insert("end", ans)

        from datetime import datetime
        t = datetime.now().strftime("%H:%M:%S")
        idx = len(history_rows)
        history_rows.append((t, q, ans, idx))
        tv.insert("", "end", values=(t, _shorten(q, 60), _shorten(ans, 60), idx))
        log(f"Hỏi: {q[:50]}... / Đáp: {ans[:50]}...")

    btn_ask.configure(command=on_ask)
    txt_q.bind("<Control-Return>", on_ask)

    log("Khởi động Q&A.")
# ---------- RUN ----------
if __name__ == "__main__":
    root = tk.Tk()
    app = StudentManagerGUI(root)
    root.mainloop()


# ===========================
# ==== AI UPGRADE START ====
# (Self-contained; safe to append without changing existing GUI code)
# ===========================

from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    import pandas as _pd_check  # type: ignore
import re, json, math, os, time
from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple, Optional

try:
    import pandas as pd
    import numpy as np
except Exception:
    pd = None
    np = None

# ---- tiny tokenizer ----
_VN_STOP = {"là","và","hoặc","nhưng","thì","của","cho","với","những","các","đã","đang","sẽ","trong","tại","từ","theo","đến","về","này","kia","ấy","đó","một"}
def _norm(s:str)->str: return re.sub(r"\s+"," ", (s or "").strip())
def _tok(s:str)->List[str]:
    s=_norm(s.lower())
    return [t for t in re.findall(r"[a-zA-Z0-9_À-ỹ]+", s) if t not in _VN_STOP]

# ---- minimal hybrid index (no sklearn dependency) ----
@dataclass
class _Doc:
    id: str
    text: str
    meta: Dict[str,Any] = field(default_factory=dict)

class _CounterVec:
    def __init__(self, txt:str):
        c={}
        for t in _tok(txt): c[t]=c.get(t,0.0)+1.0
        self.c=c
def _cos(a:_CounterVec,b:_CounterVec)->float:
    if not a.c or not b.c: return 0.0
    keys=set(a.c)|set(b.c)
    dot=sum(a.c.get(k,0.0)*b.c.get(k,0.0) for k in keys)
    na=math.sqrt(sum(v*v for v in a.c.values())); nb=math.sqrt(sum(v*v for v in b.c.values()))
    return 0.0 if na==0 or nb==0 else dot/(na*nb)

class _Index:
    def __init__(self): self.docs:List[_Doc]=[]; self.vecs:List[_CounterVec]=[]
    def add_docs(self, docs:List[_Doc]):
        self.docs.extend(docs); self.vecs.extend([_CounterVec(d.text) for d in docs])
    def clear(self): self.docs.clear(); self.vecs.clear()
    def search(self, query:str, k:int=12)->List[Tuple[_Doc,float]]:
        if not self.docs: return []
        qv=_CounterVec(query)
        sims=[_cos(qv,v) for v in self.vecs]
        order=sorted(range(len(sims)), key=lambda i:sims[i], reverse=True)[:k]
        return [(self.docs[i], float(sims[i])) for i in order]

# ---- data registry ----
class DataRegistry:
    tables: Dict[str,"Any"] = {}
    index = _Index()
    @classmethod
    def default_table(cls)->str: return next(iter(cls.tables.keys()), "scores")
    @classmethod
    def clear(cls): cls.tables.clear(); cls.index.clear()
    @classmethod
    def feed_dataframe(cls, df:"Any", name:str="scores", file:str="inmem.xlsx", sheet:str="Sheet1"):
        if pd is None: return
        cls.tables[name]=df
        docs=[]
        for i,row in df.iterrows():
            parts=[f"{col}: {row[col]}" for col in df.columns if not pd.isna(row[col])]
            if parts:
                docs.append(_Doc(f"{name}:{i}", " | ".join(map(str,parts)), {"table":name,"row":int(i),"file":file,"sheet":sheet}))
        if docs: cls.index.add_docs(docs)

# ---- tools ----
class ToolError(Exception): pass

def sql_tool(plan:Dict[str,Any]):
    if pd is None: raise ToolError("Thiếu pandas.")
    if not DataRegistry.tables: raise ToolError("Chưa có dữ liệu để truy vấn.")
    t=plan.get("table") or DataRegistry.default_table()
    if t not in DataRegistry.tables: raise ToolError(f"Bảng '{t}' không tồn tại.")
    df=DataRegistry.tables[t].copy()

    sel=plan.get("select")
    if sel:
        keep=[c for c in sel if c in df.columns]
        if not keep: raise ToolError("Cột chọn không tồn tại."); df=df[keep]
        df=df[keep]

    where=plan.get("where")
    if isinstance(where,str) and where.strip():
        try: df=df.query(where, engine="python")
        except Exception as e: raise ToolError(f"Lỗi lọc: {e}")
    elif isinstance(where,dict):
        for k,v in where.items():
            if k in df.columns: df=df[df[k]==v]

    order=plan.get("order_by") or []
    if order:
        cols=[]; asc=[]
        for spec in order:
            m=re.match(r"^\s*([^\s]+)\s*(ASC|DESC)?\s*$", str(spec), re.I)
            if not m: continue
            col,dir=m.group(1), (m.group(2) or "ASC").upper()
            if col in df.columns: cols.append(col); asc.append(dir=="ASC")
        if cols: df=df.sort_values(by=cols, ascending=asc)

    lim=plan.get("limit")
    if isinstance(lim,int) and lim>0: df=df.head(lim)
    return df.reset_index(drop=True)

def stats_tool(df:"Any", op:Dict[str,Any])->Dict[str,Any]:
    if pd is None or df is None or df.empty: return {"summary":"Không có dữ liệu."}
    typ=op.get("type","describe"); col=op.get("col")
    if typ=="topk": return {"topk": df.head(int(op.get("k",30))).to_dict(orient="records")}
    if typ=="hist" and col in df.columns:
        series=pd.to_numeric(df[col], errors="coerce").dropna()
        import numpy as _np
        hist,edges=_np.histogram(series.values, bins=int(op.get("bins",10)))
        return {"hist":hist.tolist(), "edges":[float(x) for x in edges.tolist()]}
    return {"describe": df.describe(include="all").to_dict()}

def math_tool(expr:str)->Dict[str,Any]:
    import ast
    from decimal import Decimal, getcontext
    getcontext().prec=50
    class V(ast.NodeVisitor):
        def eval(self,n):
            if isinstance(n,ast.Expression): return self.eval(n.body)
            if isinstance(n,ast.Num): return Decimal(str(n.n))
            if isinstance(n,ast.Constant) and isinstance(n.value,(int,float)): return Decimal(str(n.value))
            if isinstance(n,ast.UnaryOp):
                v=self.eval(n.operand)
                return v if isinstance(n.op,ast.UAdd) else (-v if isinstance(n.op,ast.USub) else (_err()))
            if isinstance(n,ast.BinOp):
                a=self.eval(n.left); b=self.eval(n.right)
                if isinstance(n.op,ast.Add): return a+b
                if isinstance(n.op,ast.Sub): return a-b
                if isinstance(n.op,ast.Mult): return a*b
                if isinstance(n.op,ast.Div): return a/b
                if isinstance(n.op,ast.FloorDiv): return a//b
                if isinstance(n.op,ast.Mod): return a%b
                if isinstance(n.op,ast.Pow): return a**int(b)
                _err()
            _err()
    def _err(): raise ToolError("Biểu thức không hợp lệ.")
    try:
        tree=ast.parse(expr, mode="eval")
        return {"ok":True,"expr":expr,"value":str(V().eval(tree))}
    except Exception as e:
        return {"ok":False,"expr":expr,"error":str(e)}

# ---- intent ----
def _intent(q:str)->str:
    s=q.lower()
    if any(k in s for k in ["lọc","tìm","thuộc lớp","trường"]): return "filter"
    if any(k in s for k in ["xếp hạng","top","cao nhất","sắp xếp"]): return "rank"
    if any(k in s for k in ["trung bình","thống kê","hist","phân phối","tỉ lệ"]): return "stat"
    if any(k in s for k in ["tính","=", "+","-","*","/","^"]): return "calc"
    return "qa"

# ---- AI Engine (no GUI dependency) ----
class AIEngine:
    def answer(self, user_q:str)->Dict[str,Any]:
        q=_norm(user_q)
        mode=_intent(q)

        if mode in {"filter","rank","stat"}:
            plan,_cites=self._plan(q)
            try: df=sql_tool(plan)
            except Exception as e: return self._err(q,str(e))
            view={}
            if mode=="rank":
                if "order_by" not in plan and pd is not None and "Điểm TB" in df.columns:
                    df=df.sort_values(by=["Điểm TB"], ascending=False)
                view=stats_tool(df, {"type":"topk","k": plan.get("limit",30) or 30})
            elif mode=="stat":
                num=[c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])] if pd is not None else []
                col=num[0] if num else ""
                view=stats_tool(df, {"type":"hist" if col else "describe","col":col,"bins":10})
            else:
                view={"rows": len(df)}
            return {"mode":mode,"A":"Hoàn tất.","B":"Dữ liệu lấy trực tiếp từ bảng.","C":{"plan":plan},"result": (df.to_dict(orient="records") if pd is not None else []),"citations":_cites}

        if mode=="calc":
            m=re.search(r"([\d\.\+\-\*\/\^\(\)\s%//]+)", q)
            expr=m.group(1) if m else q
            out=math_tool(expr)
            if out.get("ok"): return {"mode":"calc","A":f"Kết quả = {out['value']}","B":"Tính bằng math_tool.","C":{"expr":out["expr"]}}
            return self._err(q, out.get("error","Không tính được."))

        # qa via retrieval
        pairs=DataRegistry.index.search(q, k=12)
        if not pairs:
            return {"mode":"qa","A":"Không có dữ liệu để trả lời.","B":"Hãy nạp bảng đang xem vào DataRegistry.feed_dataframe(...).","C":{},"citations":[]}
        pairs=self._rerank(q,pairs,8)
        ans=self._synth(q,pairs)
        cites=[d.meta for d,_ in pairs]
        return {"mode":"qa","A":ans,"B":"; ".join([f"{m.get('file','?')}[{m.get('table','?')}/{m.get('sheet','?')}]#row{m.get('row','?')}" for m in cites]),"C":{"k":len(pairs)},"citations":cites}

    def _plan(self,q:str)->Tuple[Dict[str,Any],List[Dict[str,Any]]]:
        table=DataRegistry.default_table()
        where={}
        m=re.search(r"(lớp|class)\s*[:=]?\s*([A-Za-z0-9]+)", q, flags=re.I)
        if m: where["Lớp"]=m.group(2)
        m2=re.search(r"(trường)\s*[:=]?\s*(.+)", q, flags=re.I)
        if m2:
            val=re.sub(r"[,;\.\s]+$","", m2.group(2).strip())
            if val: where["Trường"]=val
        sel=["STT","ID","Họ Tên","Lớp","Toán","Lý","Hóa","Văn","Anh","Tin","Điểm TB","Trường"]
        order=[]
        if any(x in q.lower() for x in ["xếp hạng","top","cao nhất","điểm cao"]):
            subj=None
            for s in ["Toán","Lý","Hóa","Văn","Anh","Tin","Điểm TB"]:
                if s.lower() in q.lower(): subj=s; break
            order=[f"{subj or 'Điểm TB'} DESC"]
        limit=30 if "top" in q.lower() else None
        return {"op":"select","table":table,"select":sel,"where":where,"order_by":order,"limit":limit}, [{"table":table}]

    def _rerank(self,q:str,pairs:List[Tuple[_Doc,float]],k:int)->List[Tuple[_Doc,float]]:
        picked=[]; cand=pairs[:]; qt=set(_tok(q))
        while cand and len(picked)<k:
            best=None; bestv=-1e9
            for d,s in cand:
                red=max((len(set(_tok(d.text)) & set(_tok(p[0].text))) for p in picked), default=0)+1e-6
                qo=len(qt & set(_tok(d.text)))+1e-6
                v=0.7*s + 0.3*(qo/10.0) - 0.1*math.log(red)
                if v>bestv: bestv=v; best=(d,s)
            picked.append(best); cand.remove(best)
        return picked

    def _synth(self,q:str,pairs:List[Tuple[_Doc,float]])->str:
        bullets=[]
        for d,_ in pairs[:4]:
            bullets.append("- "+ (d.text[:220] + ("..." if len(d.text)>220 else "")))
        return "Hỏi: "+q+"\nTóm tắt:\n"+("\n".join(bullets))

    def _err(self,q:str,msg:str)->Dict[str,Any]:
        return {"mode":"error","A":"Không thực hiện được.","B":f"Lý do: {msg}","C":{"question":q},"citations":[]}

# Global instance + helper
AI_ENGINE = AIEngine()
def ai_answer(query:str)->Dict[str,Any]:
    return AI_ENGINE.answer(query)
# =========================
# ==== AI UPGRADE END ====
# =========================